# importar modelos.

import math
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, FileResponse
from django.utils import timezone
from django.db.models import Q
from django.core.paginator import Paginator
from django.contrib import messages
from django.conf import settings
from rest_framework import viewsets, filters
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import action
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework_simplejwt.tokens import AccessToken
from openpyxl import Workbook
from io import BytesIO
from collections import defaultdict
import requests
from datetime import datetime
import shutil
import os
from .models import Insumo, Receta, Ingrediente, MenuDiario, CalculoInsumos, Favorito, Notificacion
from .serializers import InsumoSerializer, RecetaSerializer, IngredienteSerializer, MenuDiarioSerializer
from .forms import InsumoForm, RecetaForm, IngredienteForm, MenuDiarioForm
from .permissions import IsOwnerOrReadOnly
import openpyxl
from .forms import MenuDiarioMultipleForm
from django.shortcuts import render, redirect
import json
#from .insumo_presentaciones import insumo_presentaciones  # 🔥 Acordate de importarlo
from recetas.insumo_presentaciones import insumo_presentaciones
from recetas.models import Insumo  # asegurate que esté importado



# def parsear_fecha(fecha_str):
#     """Intenta convertir fecha flexible: yyyy-mm-dd o dd/mm/yyyy."""
#     for formato in ("%Y-%m-%d", "%d/%m/%Y"):
#         try:
#             return datetime.strptime(fecha_str, formato).date()
#         except ValueError:
#             continue
#     raise ValueError(f"Formato de fecha inválido: {fecha_str}")

# --- Corregir fechas ----- #
def parsear_fecha(fecha_str):
    """Intenta convertir fecha flexible: yyyy-mm-dd o dd/mm/yyyy."""
    for formato in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(fecha_str, formato).date()
        except ValueError:
            continue
    raise ValueError(f"Formato de fecha inválido: {fecha_str}")



# --- API ViewSets ---

class RecetaViewSet(viewsets.ModelViewSet):
    queryset = Receta.objects.all()
    serializer_class = RecetaSerializer
    permission_classes = [IsAuthenticated, IsOwnerOrReadOnly]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['temporada', 'tipo_comida']
    search_fields = ['nombre']

    def perform_create(self, serializer):
        serializer.save(usuario=self.request.user)

class InsumoViewSet(viewsets.ModelViewSet):
    queryset = Insumo.objects.all()
    serializer_class = InsumoSerializer
    permission_classes = [IsAuthenticated, IsOwnerOrReadOnly]

    def perform_create(self, serializer):
        serializer.save(usuario=self.request.user)

class IngredienteViewSet(viewsets.ModelViewSet):
    queryset = Ingrediente.objects.all()
    serializer_class = IngredienteSerializer
    permission_classes = [IsAuthenticated, IsOwnerOrReadOnly]

    def perform_create(self, serializer):
        serializer.save(usuario=self.request.user)

class MenuDiarioViewSet(viewsets.ModelViewSet):
    queryset = MenuDiario.objects.all()
    serializer_class = MenuDiarioSerializer
    permission_classes = [IsAuthenticated, IsOwnerOrReadOnly]

    def perform_create(self, serializer):
        serializer.save(usuario=self.request.user)

    @action(detail=False, methods=['get'])
    def calcular_insumos(self, request):
        fecha_inicio = request.query_params.get('fecha_inicio')
        fecha_fin = request.query_params.get('fecha_fin')
        temporada = request.query_params.get('temporada', 'verano')

        try:
            fecha_inicio = datetime.strptime(fecha_inicio, '%d-%m-%Y').date()
            fecha_fin = datetime.strptime(fecha_fin, '%d-%m-%Y').date()
        except (ValueError, TypeError):
            return Response({"error": "Formato de fecha inválido. Use DD-MM-YYYY."}, status=400)

        menus = MenuDiario.objects.filter(fecha__range=[fecha_inicio, fecha_fin], temporada=temporada)
        insumos_necesarios = {}

        for menu in menus:
            receta = menu.receta
            factor = menu.comensales / receta.porciones
            ingredientes = Ingrediente.objects.filter(receta=receta)
            for ingrediente in ingredientes:
                insumo = ingrediente.insumo
                cantidad = ingrediente.cantidad * factor
                if insumo.nombre in insumos_necesarios:
                    insumos_necesarios[insumo.nombre] += cantidad
                else:
                    insumos_necesarios[insumo.nombre] = cantidad

        resultado = [
            {
                'codigo': Insumo.objects.get(nombre=nombre).codigo,
                'insumo': nombre,
                'cantidad': round(cantidad, 2),
                'unidad': Insumo.objects.get(nombre=nombre).unidad
            }
            for nombre, cantidad in insumos_necesarios.items()
        ]

        CalculoInsumos.objects.create(
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
            temporada=temporada,
            insumos=json.dumps(resultado),
            usuario=request.user
        )

        return Response(resultado)

# --- Vistas Web ---

def index(request):
    recetas = Receta.objects.all().order_by('-id')[:5]
    menus = MenuDiario.objects.all().order_by('-fecha')[:5]
    notificaciones_no_leidas = Notificacion.objects.filter(usuario=request.user, leida=False).count()
    return render(request, 'recetas/index.html', {
        'recetas': recetas,
        'menus': menus,
        'notificaciones_no_leidas': notificaciones_no_leidas,
    })



@login_required  # lo agregue 13/04 9:00am.
def agregar_insumo(request):
    if request.method == 'POST':
        form = InsumoForm(request.POST)
        if form.is_valid():
            insumo = form.save(commit=False)
            insumo.usuario = request.user
            insumo.save()
            return redirect('recetas:index')
    else:
        form = InsumoForm()
    return render(request, 'recetas/agregar_insumo.html', {'form': form})


@login_required
def agregar_receta(request):
    if request.method == 'POST':
        form = RecetaForm(request.POST, request.FILES)
        if form.is_valid():
            receta = form.save(commit=False)
            receta.usuario = request.user
            receta.save()
            return redirect('recetas:receta_list')
    else:
        form = RecetaForm()
    return render(request, 'recetas/agregar_receta.html', {'form': form})

@login_required
def agregar_ingrediente(request, receta_id):
    receta = get_object_or_404(Receta, id=receta_id)

    if request.method == 'POST':
        form = IngredienteForm(request.POST)
        if form.is_valid():
            ingrediente = form.save(commit=False)
            ingrediente.usuario = request.user
            ingrediente.receta = receta  # acá le decimos a qué receta pertenece
            ingrediente.save()
            return redirect('recetas:ver_receta_completa', receta_id=receta.id)  # o donde quieras
    else:
        form = IngredienteForm()

    return render(request, 'recetas/agregar_ingrediente.html', {'form': form, 'receta': receta})

# nuevo menú diario 

@login_required
def agregar_menu_diario_multiple(request):
    if request.method == 'POST':
        form = MenuDiarioMultipleForm(request.POST)
        if form.is_valid():
            fechas_raw = form.cleaned_data['fechas']
            temporada = form.cleaned_data['temporada']
            desayuno = form.cleaned_data['desayuno']
            merienda = form.cleaned_data['merienda']
            almuerzo = form.cleaned_data['almuerzo']
            cena = form.cleaned_data['cena']

            fechas_list = [f.strip() for f in fechas_raw.split(',') if f.strip()]
            creados = []
            duplicados = []

            for fecha_str in fechas_list:
                try:
                    fecha_obj = datetime.strptime(fecha_str, '%d-%m-%Y').date()
                    
                    existe = MenuDiario.objects.filter(fecha=fecha_obj, temporada=temporada).exists()
                    if existe:
                        duplicados.append(fecha_obj)
                        continue

                    if desayuno:
                        MenuDiario.objects.create(
                            fecha=fecha_obj,
                            temporada=temporada,
                            tipo_comida='desayuno',
                            receta=desayuno,
                            comensales=115,
                            usuario=request.user
                        )

                    if merienda:
                        MenuDiario.objects.create(
                            fecha=fecha_obj,
                            temporada=temporada,
                            tipo_comida='merienda',
                            receta=merienda,
                            comensales=115,
                            usuario=request.user
                        )

                    if almuerzo:
                        MenuDiario.objects.create(
                            fecha=fecha_obj,
                            temporada=temporada,
                            tipo_comida='almuerzo',
                            receta=almuerzo,
                            comensales=115,
                            usuario=request.user
                        )

                    if cena:
                        MenuDiario.objects.create(
                            fecha=fecha_obj,
                            temporada=temporada,
                            tipo_comida='cena',
                            receta=cena,
                            comensales=115,
                            usuario=request.user
                        )

                    creados.append(fecha_obj)

                except ValueError:
                    messages.error(request, f"Formato inválido de fecha: {fecha_str}")

            if creados:
                messages.success(request, f"Se cargaron menús para {len(creados)} días correctamente.")
            if duplicados:
                messages.warning(request, f"Estos días ya tenían menús cargados y se omitieron: {', '.join([str(d) for d in duplicados])}")

            return redirect('recetas:agregar_menu_diario_multiple')

    else:
        form = MenuDiarioMultipleForm()

    return render(request, 'recetas/agregar_menu_diario_multiple.html', {'form': form})

@login_required
def agregar_menu_diario(request):
    if request.method == 'POST':
        form = MenuDiarioForm(request.POST)
        if form.is_valid():
            menu = form.save(commit=False)
            menu.usuario = request.user  # 🎯 Guarda el usuario que está logueado
            menu.save()
            return redirect('recetas:agregar_menu_diario')
    else:
        form = MenuDiarioForm()
    return render(request, 'recetas/agregar_menu_diario.html', {'form': form})


@login_required
def calcular_insumos_web(request):
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    temporada = request.GET.get('temporada')

    insumos = []
    menus = []

    if fecha_inicio and fecha_fin and temporada:
        try:
            fecha_inicio_obj = parsear_fecha(fecha_inicio)
            fecha_fin_obj = parsear_fecha(fecha_fin)

            filtros = {
                'fecha__range': (fecha_inicio_obj, fecha_fin_obj),
                'temporada': temporada,
            }

            menus = MenuDiario.objects.filter(**filtros).select_related('receta').prefetch_related('receta__ingredientes__insumo')

            insumo_totales = defaultdict(lambda: {'nombre': '', 'unidad': '', 'cantidad': 0})

            for menu in menus:
                receta = menu.receta
                comensales = menu.comensales

                for ingrediente in receta.ingredientes.all():
                    insumo = ingrediente.insumo
                    cantidad_total = ingrediente.cantidad * comensales

                    codigo = int(insumo.codigo) if isinstance(insumo.codigo, str) else insumo.codigo
                    presentacion = insumo_presentaciones.get(codigo, 1)

                    cantidad_final = math.ceil(cantidad_total / presentacion)

                    insumo_totales[codigo]['nombre'] = insumo.nombre
                    insumo_totales[codigo]['unidad'] = insumo.unidad
                    insumo_totales[codigo]['cantidad'] += cantidad_final

                    print(f"🔍 Procesando Insumo Código {codigo} - {insumo.nombre}")
                    print(f"Cantidad bruta: {cantidad_total}")
                    print(f"Presentación encontrada: {presentacion}")
                    print(f"Cantidad final (paquetes redondeados): {cantidad_final}")

            insumos = [
                {'codigo': codigo, 'insumo': datos['nombre'], 'unidad': datos['unidad'], 'cantidad': datos['cantidad']}
                for codigo, datos in insumo_totales.items()
            ]

            insumos = sorted(insumos, key=lambda x: int(x['codigo']))

            CalculoInsumos.objects.create(
                fecha_inicio=fecha_inicio_obj,
                fecha_fin=fecha_fin_obj,
                temporada=temporada,
                insumos=json.dumps(insumos),
                usuario=request.user
            )

        except Exception as e:
            print(f"Error en el cálculo de insumos: {e}")

    contexto = {
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'temporada': temporada,
        'insumos': insumos,
        'menus': menus,
    }

    return render(request, 'recetas/calcular_insumos.html', contexto)



@login_required
def exportar_insumos_excel(request):
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    temporada = request.GET.get('temporada')

    if not (fecha_inicio and fecha_fin and temporada):
        return HttpResponse("Faltan parámetros para exportar.", status=400)

    try:
        fecha_inicio_obj = parsear_fecha(fecha_inicio)
        fecha_fin_obj = parsear_fecha(fecha_fin)

        filtros = {
            'fecha__range': (fecha_inicio_obj, fecha_fin_obj),
            'temporada': temporada,
        }

        menus = MenuDiario.objects.filter(**filtros).select_related('receta').prefetch_related('receta__ingredientes__insumo')

        insumo_totales = defaultdict(lambda: {'nombre': '', 'unidad': '', 'cantidad': 0})

        for menu in menus:
            receta = menu.receta
            comensales = menu.comensales

            for ingrediente in receta.ingredientes.all():
                insumo = ingrediente.insumo
                cantidad_total = ingrediente.cantidad * comensales
                codigo = int(insumo.codigo) if isinstance(insumo.codigo, str) else insumo.codigo
                presentacion = insumo_presentaciones.get(codigo, 1)
                cantidad_final = math.ceil(cantidad_total / presentacion)

                insumo_totales[codigo]['nombre'] = insumo.nombre
                insumo_totales[codigo]['unidad'] = insumo.unidad
                insumo_totales[codigo]['cantidad'] += cantidad_final

    except Exception as e:
        return HttpResponse(f"Error al calcular insumos: {e}", status=500)

    # Creamos el Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Insumos Necesarios"

    ws.append(['Código', 'Insumo', 'Cantidad'])

    codigos_con_salto = {
        3: 2098,  # Acelga después de Atún
        1: 999,   # Aceite después de Cebolla de verdeo
        70: 7778,  # Pan después de algún código anterior (puede que sea 788 o lo que vos decidas)
    }

    ultimo_codigo = None
    for codigo in insumo_presentaciones.keys():
        if ultimo_codigo and codigo in codigos_con_salto and codigos_con_salto[codigo] == ultimo_codigo:
            ws.append([])  # Fila vacía

        datos = insumo_totales.get(codigo)
        if datos:
            cantidad = datos['cantidad']
            nombre = datos['nombre']
        else:
            cantidad = 0
            try:
                nombre = Insumo.objects.get(codigo=codigo).nombre
            except Insumo.DoesNotExist:
                nombre = '(desconocido)'

        ws.append([codigo, nombre, cantidad])
        ultimo_codigo = codigo

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"insumos_{fecha_inicio}_a_{fecha_fin}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response




@login_required
def historico_calculos(request):
    calculos = CalculoInsumos.objects.all().order_by('-fecha_calculo')
    for calculo in calculos:
        try:
            calculo.insumos = json.loads(calculo.insumos)
        except Exception:
            calculo.insumos = []

    return render(request, 'recetas/historico_calculos.html', {'calculos': calculos})



@login_required
def receta_list(request):
    query = request.GET.get('q', '')
    temporada = request.GET.get('temporada', '')
    tipo_comida = request.GET.get('tipo_comida', '')
    recetas = Receta.objects.all()
    if query:
        recetas = recetas.filter(nombre__icontains=query)
    if temporada:
        recetas = recetas.filter(temporada=temporada)
    if tipo_comida:
        recetas = recetas.filter(tipo_comida=tipo_comida)
    return render(request, 'recetas/receta_list.html', {
        'recetas': recetas,
        'query': query,
        'temporada': temporada,
        'tipo_comida': tipo_comida,
    })

@login_required
def receta_edit(request, receta_id):
    receta = get_object_or_404(Receta, id=receta_id)
    if request.method == 'POST':
        form = RecetaForm(request.POST, request.FILES, instance=receta)
        if form.is_valid():
            form.save()
            return redirect('recetas:gestionar_recetas')
    else:
        form = RecetaForm(instance=receta)
    return render(request, 'recetas/receta_form.html', {'form': form})

@login_required
def receta_delete(request, receta_id):
    receta = get_object_or_404(Receta, id=receta_id)

    if request.method == 'POST':
        receta.delete()
        return redirect('recetas:gestionar_recetas')

    return render(request, 'recetas/confirmar_eliminar_receta.html', {'receta': receta})

def ver_receta_completa(request, receta_id):
    receta = get_object_or_404(Receta, id=receta_id)
    ingredientes = receta.ingredientes.all()  # <- obtenemos ingredientes relacionados
    return render(request, 'recetas/ver_receta_completa.html', {
        'receta': receta,
        'ingredientes': ingredientes,  # <- los mandamos al template
    })

    
@login_required
def toggle_favorito(request, pk):
    receta = get_object_or_404(Receta, pk=pk)
    favorito, created = Favorito.objects.get_or_create(usuario=request.user, receta=receta)
    if created:
        Notificacion.objects.create(
            usuario=request.user,
            mensaje=f"Has añadido '{receta.nombre}' a tus favoritos."
        )
    else:
        favorito.delete()
        Notificacion.objects.create(
            usuario=request.user,
            mensaje=f"Has eliminado '{receta.nombre}' de tus favoritos."
        )
    return redirect('recetas:ver_receta_completa', pk=pk)

@login_required
def favoritos_list(request):
    favoritos = Favorito.objects.filter(usuario=request.user)
    return render(request, 'recetas/favoritos_list.html', {'favoritos': favoritos})

@login_required
def agregar_comentario(request, pk):
    receta = get_object_or_404(Receta, pk=pk)
    if request.method == 'POST':
        form = ComentarioForm(request.POST)
        if form.is_valid():
            comentario = form.save(commit=False)
            comentario.usuario = request.user
            comentario.receta = receta
            comentario.save()
            return redirect('recetas:ver_receta_completa', pk=pk)
    else:
        form = ComentarioForm()
    return render(request, 'recetas/agregar_comentario.html', {'form': form, 'receta': receta})

@login_required
def responder_comentario(request, comentario_id):
    comentario = get_object_or_404(Comentario, id=comentario_id)
    if request.method == 'POST':
        form = ComentarioForm(request.POST)
        if form.is_valid():
            respuesta = form.save(commit=False)
            respuesta.usuario = request.user
            respuesta.receta = comentario.receta
            respuesta.padre = comentario
            respuesta.save()
            return redirect('recetas:ver_receta_completa', pk=comentario.receta.id)
    else:
        form = ComentarioForm()
    return render(request, 'recetas/responder_comentario.html', {'form': form, 'comentario': comentario})

@login_required
def eliminar_comentario(request, comentario_id):
    comentario = get_object_or_404(Comentario, id=comentario_id, usuario=request.user)
    receta_id = comentario.receta.id
    if request.method == 'POST':
        comentario.delete()
        return redirect('recetas:ver_receta_completa', pk=receta_id)
    return render(request, 'recetas/eliminar_comentario.html', {'comentario': comentario})

@login_required
def notificaciones_list(request):
    notificaciones = Notificacion.objects.filter(usuario=request.user).order_by('-fecha')
    if request.method == 'POST':
        notificaciones.update(leida=True)  # Marcar todas como leídas
    return render(request, 'recetas/notificaciones_list.html', {'notificaciones': notificaciones})

@login_required
def exportar_receta_pdf(request, pk):
    receta = get_object_or_404(Receta, pk=pk)
    ingredientes = Ingrediente.objects.filter(receta=receta)

    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)

    # Encabezado
    p.setFont("Helvetica-Bold", 10)
    p.drawString(50, 780, "Gestión de Recetas")
    p.line(50, 775, 550, 775)

    # Título
    p.setFont("Helvetica-Bold", 16)
    p.drawString(100, 740, f"Receta: {receta.nombre}")

    # Detalles
    p.setFont("Helvetica", 12)
    p.drawString(100, 710, f"Código: {receta.codigo}")
    p.drawString(100, 690, f"Temporada: {receta.temporada}")
    p.drawString(100, 670, f"Tipo de Comida: {receta.tipo_comida}")
    p.drawString(100, 650, f"Porciones: {receta.porciones}")

    # Ingredientes
    p.setFont("Helvetica-Bold", 12)
    p.drawString(100, 620, "Ingredientes:")
    p.setFont("Helvetica", 12)
    y = 600
    for ingrediente in ingredientes:
        p.drawString(120, y, f"- {ingrediente.insumo.nombre}: {ingrediente.cantidad} {ingrediente.insumo.unidad}")
        y -= 20

    # Pie de página
    p.setFont("Helvetica", 8)
    p.drawString(50, 30, f"Generado el {timezone.now().strftime('%d-%m-%Y %H:%M:%S')}")
    p.drawString(450, 30, "Página 1")

    p.showPage()
    p.save()

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="receta_{receta.codigo}.pdf"'
    return response

def listar_urls(request):
    urls = []

    def extract_urls(url_patterns, prefix=''):
        for pattern in url_patterns:
            if hasattr(pattern, 'url_patterns'):
                extract_urls(pattern.url_patterns, prefix + str(pattern.pattern))
            else:
                full_url = prefix + str(pattern.pattern)
                urls.append({
                    'url': full_url,
                    'name': pattern.name if pattern.name else 'Sin nombre',
                    'view': str(pattern.callback) if pattern.callback else 'Sin vista'
                })

    extract_urls(get_resolver().url_patterns)
    return render(request, 'recetas/listar_urls.html', {'urls': urls})



from django.shortcuts import get_object_or_404, redirect
from django.shortcuts import render, get_object_or_404, redirect  # <-- agregá redirect acá si no lo tenías

@login_required
def borrar_calculo(request, calculo_id):
    try:
        calculo = CalculoInsumos.objects.get(id=calculo_id)
        calculo.delete()
        messages.success(request, "Cálculo eliminado exitosamente. 🗑️")
    except CalculoInsumos.DoesNotExist:
        messages.error(request, "El cálculo no existe.")
    return redirect('recetas:historico_calculos')


@login_required
def test_borrar_calculo(request):
    return HttpResponse("Funciona borrar_calculo")

@login_required
def gestionar_recetas(request):
    query = request.GET.get('q', '')
    recetas = Receta.objects.all()

    if query:
        recetas = recetas.filter(
            Q(nombre__icontains=query) |
            Q(codigo__icontains=query)
        )

    return render(request, 'recetas/gestionar_recetas.html', {'recetas': recetas, 'query': query})

@login_required
def importar_insumos(request):
    if request.method == 'POST' and request.FILES.get('archivo'):
        archivo = request.FILES['archivo']
        try:
            wb = openpyxl.load_workbook(archivo)
            hoja = wb.active

            # Empezamos desde la segunda fila (asumiendo que la primera tiene encabezados)
            for fila in hoja.iter_rows(min_row=2, values_only=True):
                codigo, nombre, unidad, presentacion = fila

                if codigo and nombre and unidad and presentacion:
                    insumo, creado = Insumo.objects.update_or_create(
                        codigo=codigo,
                        defaults={
                            'nombre': nombre,
                            'unidad': unidad,
                            'presentacion': presentacion,
                            'usuario': request.user
                        }
                    )
            messages.success(request, '¡Insumos importados y actualizados exitosamente!')
            return redirect('recetas:importar_insumos')

        except Exception as e:
            messages.error(request, f"Error al procesar el archivo: {e}")
            return redirect('recetas:importar_insumos')

    return render(request, 'recetas/importar_insumos.html')

@login_required
def importar_recetas(request):
    if request.method == 'POST' and request.FILES.get('archivo'):
        archivo = request.FILES['archivo']
        try:
            wb = openpyxl.load_workbook(archivo)
            hoja = wb.active

            for fila in hoja.iter_rows(min_row=2, values_only=True):
                codigo, nombre, temporada, tipo_comida, porciones = fila

                if codigo and nombre and temporada and tipo_comida:
                    receta, creada = Receta.objects.update_or_create(
                        codigo=codigo,
                        defaults={
                            'nombre': nombre,
                            'temporada': temporada,
                            'tipo_comida': tipo_comida,
                            'porciones': porciones if porciones else 1,
                            'usuario': request.user
                        }
                    )

            messages.success(request, '¡Recetas importadas y actualizadas exitosamente!')
            return redirect('recetas:importar_recetas')

        except Exception as e:
            messages.error(request, f"Error al procesar el archivo: {e}")
            return redirect('recetas:importar_recetas')

    return render(request, 'recetas/importar_recetas.html')

@login_required
def importar_ingredientes(request):
    if request.method == 'POST' and request.FILES.get('archivo'):
        archivo = request.FILES['archivo']
        try:
            wb = openpyxl.load_workbook(archivo)
            hoja = wb.active

            errores = []  # Para listar errores

            for fila_num, fila in enumerate(hoja.iter_rows(min_row=2, values_only=True), start=2):
                codigo_receta, codigo_insumo, cantidad = fila

                # Validamos que existan datos y que cantidad sea positiva (> 0)
                if not (codigo_receta and codigo_insumo) or cantidad in (None, '', 0):
                    errores.append(f"Fila {fila_num}: datos incompletos o cantidad inválida (Receta: {codigo_receta}, Insumo: {codigo_insumo}, Cantidad: {cantidad})")
                    continue  # No procesamos esta fila

                try:
                    receta = Receta.objects.get(codigo=codigo_receta)
                    insumo = Insumo.objects.get(codigo=codigo_insumo)

                    ingrediente, creado = Ingrediente.objects.update_or_create(
                        receta=receta,
                        insumo=insumo,
                        defaults={
                            'cantidad': cantidad,
                            'usuario': request.user
                        }
                    )
                except Receta.DoesNotExist:
                    errores.append(f"Fila {fila_num}: Receta '{codigo_receta}' no encontrada.")
                except Insumo.DoesNotExist:
                    errores.append(f"Fila {fila_num}: Insumo '{codigo_insumo}' no encontrado.")
                except Exception as e:
                    errores.append(f"Fila {fila_num}: Error general {e}")

            if errores:
                messages.warning(request, f"Importación terminada con errores:\n" + "\n".join(errores))
            else:
                messages.success(request, '¡Ingredientes importados exitosamente!')

            return redirect('recetas:importar_ingredientes')

        except Exception as e:
            messages.error(request, f"Error al procesar el archivo: {e}")
            return redirect('recetas:importar_ingredientes')

    return render(request, 'recetas/importar_ingredientes.html')



#Backup de la base de datos.
@login_required
def backup_bd(request):
    ruta_original = settings.BASE_DIR / 'db.sqlite3'
    ruta_backup = settings.BASE_DIR / 'backup_db.sqlite3'

    # Copiamos el archivo
    shutil.copy(ruta_original, ruta_backup)

    # Devolvemos el archivo como descarga
    response = FileResponse(open(ruta_backup, 'rb'), as_attachment=True, filename='backup_db.sqlite3')

    return response

@login_required
def ingredientes_list(request):
    buscar = request.GET.get('buscar', '')

    recetas = Receta.objects.prefetch_related('ingredientes__insumo')

    if buscar:
        recetas = recetas.filter(
            Q(nombre__icontains=buscar) | Q(codigo__icontains=buscar)
        )

    ingredientes_por_receta = []
    for receta in recetas:
        ingredientes = receta.ingredientes.all()
        if ingredientes:
            ingredientes_por_receta.append((receta, ingredientes))

    paginator = Paginator(ingredientes_por_receta, 10)  # Paginamos
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'buscar': buscar,  # Esto lo mandamos para que en el form siga mostrando lo buscado
    }

    return render(request, 'recetas/ingredientes_list.html', context)

    
    

@login_required
def editar_ingrediente(request, ingrediente_id):
    ingrediente = get_object_or_404(Ingrediente, id=ingrediente_id)

    if request.method == 'POST':
        form = IngredienteForm(request.POST, instance=ingrediente)
        if form.is_valid():
            form.save()
            messages.success(request, 'Ingrediente actualizado correctamente.')
            return redirect('recetas:ingredientes_list')
    else:
        form = IngredienteForm(instance=ingrediente)
    return render(request, 'recetas/editar_ingrediente.html', {'form': form, 'ingrediente': ingrediente})


@login_required
def eliminar_ingrediente(request, ingrediente_id):
    ingrediente = get_object_or_404(Ingrediente, id=ingrediente_id)

    if request.method == 'POST':
        ingrediente.delete()
        messages.success(request, 'Ingrediente eliminado correctamente.')
        return redirect('recetas:ingredientes_list')

    return render(request, 'recetas/eliminar_ingrediente.html', {'ingrediente': ingrediente})

# recetas/views.py


from .forms import MenuDiarioMultipleForm

@login_required
def agregar_menu_multiple(request):
    if request.method == 'POST':
        form = MenuDiarioMultipleForm(request.POST)
        if form.is_valid():
            fecha = form.cleaned_data['fecha']

            registros = [
                ('desayuno', form.cleaned_data.get('desayuno'), form.cleaned_data.get('comensales_desayuno')),
                ('merienda', form.cleaned_data.get('merienda'), form.cleaned_data.get('comensales_merienda')),
                ('almuerzo', form.cleaned_data.get('almuerzo'), form.cleaned_data.get('comensales_almuerzo')),
                ('cena', form.cleaned_data.get('cena'), form.cleaned_data.get('comensales_cena')),
            ]

            for tipo_comida, receta, comensales in registros:
                if receta and comensales and comensales > 0:
                    MenuDiario.objects.update_or_create(
                        fecha=fecha,
                        tipo_comida=tipo_comida,
                        temporada=receta.temporada,
                        defaults={
                            'receta': receta,
                            'comensales': comensales,
                            'usuario': request.user
                        }
                    )

            messages.success(request, '¡Menú diario cargado correctamente!')
            return redirect('recetas:agregar_menu_multiple')
    else:
        form = MenuDiarioMultipleForm()

    menus = MenuDiario.objects.filter(usuario=request.user).order_by('-fecha', 'tipo_comida')

    contexto = {
        'form': form,
        'menus': menus
    }

    return render(request, 'recetas/agregar_menu_multiple.html', contexto)



# @login_required
# def ver_calculo(request, calculo_id):
#     try:
#         calculo = CalculoInsumos.objects.get(id=calculo_id)
#         insumos_guardados = {int(i['codigo']): i for i in json.loads(calculo.insumos)}
#     except CalculoInsumos.DoesNotExist:
#         messages.error(request, "El cálculo seleccionado no existe.")
#         return redirect('recetas:historico_calculos')
#     except Exception as e:
#         messages.error(request, f"Error al cargar el cálculo: {e}")
#         return redirect('recetas:historico_calculos')

#     # Creamos la lista ordenada incluyendo ceros
#     insumos_ordenados = []
#     for codigo in insumo_presentaciones.keys():
#         if codigo in insumos_guardados:
#             item = insumos_guardados[codigo]
#             cantidad = item.get('cantidad', 0)
#         else:
#             cantidad = 0
#             item = {
#                 'codigo': codigo,
#                 'insumo': f'(sin nombre)',  # opcional
#                 'unidad': '',  # no mostramos igual
#             }

#         insumos_ordenados.append({
#             'codigo': codigo,
#             'insumo': item['insumo'],
#             'cantidad': cantidad
#         })

#     return render(request, 'recetas/ver_calculo.html', {
#         'calculo': calculo,
#         'insumos': insumos_ordenados,
#     })
@login_required
def ver_calculo(request, calculo_id):
    try:
        calculo = CalculoInsumos.objects.get(id=calculo_id)
        insumos_guardados = json.loads(calculo.insumos)
    except CalculoInsumos.DoesNotExist:
        messages.error(request, "El cálculo seleccionado no existe.")
        return redirect('recetas:historico_calculos')
    except Exception as e:
        messages.error(request, f"Error al cargar el cálculo: {e}")
        return redirect('recetas:historico_calculos')

    insumos_ordenados = []

    for codigo in insumo_presentaciones.keys():
        # Buscar el insumo en los guardados
        item = next((i for i in insumos_guardados if int(i.get('codigo', 0)) == codigo), None)

        if item:
            cantidad = item.get('cantidad', 0)
            nombre = item.get('insumo', '')
        else:
            cantidad = 0
            nombre = ''

        if not nombre:
            try:
                nombre = Insumo.objects.get(codigo=codigo).nombre
            except Insumo.DoesNotExist:
                nombre = "(sin nombre)"

        insumos_ordenados.append({
            'codigo': codigo,
            'insumo': nombre,
            'cantidad': cantidad
        })

    return render(request, 'recetas/ver_calculo.html', {
        'calculo': calculo,
        'insumos': insumos_ordenados
    })


@login_required
def borrar_menu_diario(request, menu_id):
    try:
        menu = MenuDiario.objects.get(id=menu_id)
        menu.delete()
        messages.success(request, "El menú diario fue borrado correctamente.")
    except MenuDiario.DoesNotExist:
        messages.error(request, "El menú diario no existe.")
    return redirect('recetas:agregar_menu_multiple')



# insumos_ordenados = []
# for codigo in insumo_presentaciones.keys():
#     if codigo in insumos_guardados:
#         item = insumos_guardados[codigo]
#         cantidad = item.get('cantidad', 0)
#         nombre = item.get('insumo', '')
#     else:
#         cantidad = 0
#         try:
#             nombre = Insumo.objects.get(codigo=codigo).nombre
#         except Insumo.DoesNotExist:
#             nombre = '(desconocido)'

#     insumos_ordenados.append({
#         'codigo': codigo,
#         'insumo': nombre,
#         'cantidad': cantidad
#     })
